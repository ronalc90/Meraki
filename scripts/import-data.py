"""
Script de importación de datos desde Excel a Supabase.
Importa: productos (Costos), inventario, y pedidos de abril.

Uso:
  pip install openpyxl supabase
  python scripts/import-data.py

Configurar las variables SUPABASE_URL y SUPABASE_KEY antes de ejecutar.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalar: pip install openpyxl")
    sys.exit(1)

# Rutas de los archivos Excel
EXCEL_DIR = Path.home() / "Downloads"
ABRIL_FILE = EXCEL_DIR / "Abril2026.xlsx"
INVENTARIO_FILE = EXCEL_DIR / "Inventario2025 v2.xlsm"

OUTPUT_DIR = Path(__file__).parent / "import-output"
OUTPUT_DIR.mkdir(exist_ok=True)


def extract_products():
    """Extraer productos de la hoja 'Costos' de Abril2026.xlsx"""
    wb = openpyxl.load_workbook(ABRIL_FILE, data_only=True)
    ws = wb['Costos']
    products = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        raw_code = row[0]
        try:
            code_num = str(int(float(raw_code)))
        except (ValueError, TypeError):
            code_num = str(raw_code).strip()
        name = str(row[1] or '').strip()
        cost = int(float(row[2] or 0))

        if not name:
            continue

        # Determinar categoría
        name_lower = name.lower()
        if 'maxisaco' in name_lower or 'max' in name_lower:
            category = 'Maxisaco'
            code = 'MAX'
        elif 'bota' in name_lower:
            category = 'Pantuflas'
            code = 'PANT'
        elif 'pocillo' in name_lower or 'set' in name_lower:
            category = 'Pocillo'
            code = 'POC'
        else:
            category = 'Pantuflas'
            code = 'PANT'

        products.append({
            'code': f"{code_num}",
            'name': name,
            'cost': cost,
            'category': category,
            'active': True,
        })

    wb.close()
    print(f"  Productos extraídos: {len(products)}")
    return products


def extract_orders():
    """Extraer pedidos de las hojas diarias de Abril2026.xlsx"""
    wb = openpyxl.load_workbook(ABRIL_FILE, data_only=True)
    orders = []

    for day in range(1, 32):
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, values_only=True):
            # Columnas: A=ID, B=Cliente, C=Celular, D=Dirección, E=Complemento,
            # F=Ref, G=Detalle, H=Comentario, I=Valor, J=Efectivo Bogo,
            # K=Caja, L=Transferencia, M=Costo, N=Entrega, O=Vendedor,
            # P=Estado, Q=Complemento estado, R=Es cambio
            if row[1] is None:  # Sin cliente = fila vacía
                continue

            order_code = str(int(row[0])) if row[0] else ''
            client_name = str(row[1] or '').strip()
            phone = str(int(row[2])) if row[2] else ''
            address = str(row[3] or '').strip()
            complement = str(row[4] or '').strip()
            product_ref = str(row[5] or '').strip()
            detail = str(row[6] or '').strip()
            comment = str(row[7] or '').strip()

            value = int(row[8] or 0) if isinstance(row[8], (int, float)) else 0
            cash_bogo = int(row[9] or 0) if isinstance(row[9], (int, float)) else 0
            cash = int(row[10] or 0) if isinstance(row[10], (int, float)) else 0
            transfer = int(row[11] or 0) if isinstance(row[11], (int, float)) else 0
            cost = int(row[12] or 0) if isinstance(row[12], (int, float)) else 0

            delivery_type = str(row[13] or '').strip() if len(row) > 13 else ''
            vendor = str(row[14] or '').strip() if len(row) > 14 else ''
            status = str(row[15] or 'Confirmado').strip() if len(row) > 15 else 'Confirmado'
            status_comp = str(row[16] or '').strip() if len(row) > 16 else ''
            is_exchange = str(row[17] or '').strip().lower() == 'si' if len(row) > 17 else False

            orders.append({
                'order_code': order_code,
                'client_name': client_name,
                'phone': phone,
                'city': 'Bogotá',
                'address': address,
                'complement': complement,
                'product_ref': product_ref,
                'detail': detail,
                'comment': comment,
                'value_to_collect': value,
                'payment_cash_bogo': cash_bogo,
                'payment_cash': cash,
                'payment_transfer': transfer,
                'product_cost': cost,
                'delivery_type': delivery_type,
                'vendor': vendor,
                'delivery_status': status if status else 'Confirmado',
                'status_complement': status_comp,
                'is_exchange': is_exchange,
                'order_date': f'2026-04-{day:02d}',
            })

    wb.close()
    print(f"  Pedidos extraídos: {len(orders)}")
    return orders


def extract_inventory():
    """Extraer inventario de Inventario2025 v2.xlsm"""
    wb = openpyxl.load_workbook(INVENTARIO_FILE, read_only=True, data_only=True)
    items = []

    for sheet_name in ['Verificado Canastas', 'Malo Canastas']:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        status = 'Bueno' if 'Verificado' in sheet_name else 'Malo'
        verified = 'Verificado' in sheet_name

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8 or row[0] is None:
                continue

            basket = str(row[0] or '').strip()
            product_id = str(row[1] or '').strip()
            category = str(row[2] or 'Pantuflas').strip()
            tipo = str(row[3] or 'Adulto').strip()
            reference = int(float(row[4] or 0)) if isinstance(row[4], (int, float)) else 0
            model = str(row[5] or '').strip()
            color = str(row[6] or '').strip()
            size = str(row[7] or '').strip()
            quantity = int(float(row[8] or 1)) if len(row) > 8 and isinstance(row[8], (int, float)) else 1

            items.append({
                'basket_location': basket,
                'product_id': product_id,
                'category': category,
                'type': tipo,
                'reference': reference,
                'model': model,
                'color': color,
                'size': size,
                'quantity': quantity,
                'status': status,
                'observations': '',
                'verified': verified,
            })

    wb.close()
    print(f"  Items de inventario extraídos: {len(items)}")
    return items


def generate_sql(products, orders, inventory):
    """Genera archivo SQL para importar directamente en Supabase"""
    lines = ["-- Datos importados desde Excel\n"]

    # Products
    lines.append("-- PRODUCTOS")
    for p in products:
        name = p['name'].replace("'", "''")
        lines.append(
            f"INSERT INTO products (code, name, cost, category, active) "
            f"VALUES ('{p['code']}', '{name}', {p['cost']}, '{p['category']}', true);"
        )

    lines.append("\n-- PEDIDOS")
    for o in orders:
        client = o['client_name'].replace("'", "''")
        addr = o['address'].replace("'", "''")
        comp = o['complement'].replace("'", "''")
        det = o['detail'].replace("'", "''")
        com = o['comment'].replace("'", "''")
        lines.append(
            f"INSERT INTO orders (order_code, client_name, phone, city, address, complement, "
            f"product_ref, detail, comment, value_to_collect, payment_cash_bogo, payment_cash, "
            f"payment_transfer, product_cost, delivery_type, vendor, delivery_status, "
            f"status_complement, is_exchange, order_date) VALUES ("
            f"'{o['order_code']}', '{client}', '{o['phone']}', '{o['city']}', '{addr}', "
            f"'{comp}', '{o['product_ref']}', '{det}', '{com}', {o['value_to_collect']}, "
            f"{o['payment_cash_bogo']}, {o['payment_cash']}, {o['payment_transfer']}, "
            f"{o['product_cost']}, '{o['delivery_type']}', '{o['vendor']}', "
            f"'{o['delivery_status']}', '{o['status_complement']}', "
            f"{str(o['is_exchange']).lower()}, '{o['order_date']}');"
        )

    lines.append("\n-- INVENTARIO")
    for i in inventory:
        model = i['model'].replace("'", "''")
        lines.append(
            f"INSERT INTO inventory (basket_location, product_id, category, type, reference, "
            f"model, color, size, quantity, status, verified) VALUES ("
            f"'{i['basket_location']}', '{i['product_id']}', '{i['category']}', "
            f"'{i['type']}', {i['reference']}, '{model}', '{i['color']}', "
            f"'{i['size']}', {i['quantity']}, '{i['status']}', {str(i['verified']).lower()});"
        )

    return '\n'.join(lines)


if __name__ == '__main__':
    print("Importando datos desde Excel...\n")

    print("1. Extrayendo productos...")
    products = extract_products()

    print("2. Extrayendo pedidos...")
    orders = extract_orders()

    print("3. Extrayendo inventario...")
    inventory = extract_inventory()

    # Guardar como JSON
    json_file = OUTPUT_DIR / "data.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump({
            'products': products,
            'orders': orders,
            'inventory': inventory,
        }, f, ensure_ascii=False, indent=2)
    print(f"\nJSON guardado en: {json_file}")

    # Guardar como SQL
    sql_content = generate_sql(products, orders, inventory)
    sql_file = OUTPUT_DIR / "import-data.sql"
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"SQL guardado en: {sql_file}")

    print(f"\nResumen:")
    print(f"  - {len(products)} productos")
    print(f"  - {len(orders)} pedidos")
    print(f"  - {len(inventory)} items de inventario")
    print(f"\nPara importar en Supabase:")
    print(f"  1. Ve a https://app.supabase.com → tu proyecto → SQL Editor")
    print(f"  2. Primero ejecuta: supabase-schema.sql (crear tablas)")
    print(f"  3. Luego ejecuta: {sql_file} (importar datos)")
